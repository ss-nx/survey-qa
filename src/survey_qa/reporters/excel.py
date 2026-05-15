"""Excel report writer.

Produces a .xlsx file with three sheets:
  Summary   — headline stats and metadata
  Findings  — every Finding with colour-coded severity
  Questions — all XML questions parsed from the survey
"""

from __future__ import annotations

from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from ..core.models import Finding, SurveyModel

# ── Colour palette ─────────────────────────────────────────────────────────────

_HEADER_FILL = PatternFill("solid", fgColor="1F4E79")   # dark blue
_ERROR_FILL = PatternFill("solid", fgColor="FCE4D6")    # light red
_WARNING_FILL = PatternFill("solid", fgColor="FFEB9C")  # light yellow
_INFO_FILL = PatternFill("solid", fgColor="DDEBF7")     # light blue
_ALT_FILL = PatternFill("solid", fgColor="F2F2F2")      # light grey alternate row

_HEADER_FONT = Font(bold=True, color="FFFFFF")
_BOLD = Font(bold=True)


def _header_row(ws, values: list[str]) -> None:
    """Write a styled header row."""
    for col, value in enumerate(values, start=1):
        cell = ws.cell(row=1, column=col, value=value)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center")


def _autofit(ws) -> None:
    """Set column widths based on content (capped at 80)."""
    for col in ws.columns:
        max_len = max((len(str(c.value or "")) for c in col), default=8)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 4, 80)


def _severity_fill(severity: str) -> PatternFill:
    return {
        "error": _ERROR_FILL,
        "warning": _WARNING_FILL,
        "info": _INFO_FILL,
    }.get(severity, _INFO_FILL)


# ── Sheet builders ─────────────────────────────────────────────────────────────


def _build_summary(ws, survey: SurveyModel, findings: list[Finding]) -> None:
    ws.title = "Summary"

    errors = sum(1 for f in findings if f.severity == "error")
    warnings = sum(1 for f in findings if f.severity == "warning")
    infos = sum(1 for f in findings if f.severity == "info")

    rows = [
        ("Report generated", datetime.now().strftime("%Y-%m-%d %H:%M")),
        ("Survey label", survey.survey_label),
        ("Total elements", len(survey.elements)),
        ("Total questions", len(survey.questions())),
        ("", ""),
        ("Errors", errors),
        ("Warnings", warnings),
        ("Info", infos),
        ("Total findings", len(findings)),
    ]

    for r_idx, (label, value) in enumerate(rows, start=1):
        ws.cell(row=r_idx, column=1, value=label).font = _BOLD
        ws.cell(row=r_idx, column=2, value=value)

    # Highlight the counts
    severity_rows = {"Errors": _ERROR_FILL, "Warnings": _WARNING_FILL, "Info": _INFO_FILL}
    for r_idx, (label, _) in enumerate(rows, start=1):
        if label in severity_rows:
            ws.cell(row=r_idx, column=1).fill = severity_rows[label]
            ws.cell(row=r_idx, column=2).fill = severity_rows[label]

    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 30


def _build_findings(ws, findings: list[Finding]) -> None:
    ws.title = "Findings"

    headers = ["Check ID", "Severity", "Question Label", "Message", "Detail"]
    _header_row(ws, headers)

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"

    # Sort: errors first, then warnings, then info
    severity_order = {"error": 0, "warning": 1, "info": 2}
    sorted_findings = sorted(findings, key=lambda f: (severity_order.get(f.severity, 3), f.question_label))

    for r_idx, f in enumerate(sorted_findings, start=2):
        fill = _severity_fill(f.severity)
        values = [f.check_id, f.severity.upper(), f.question_label, f.message, f.detail]
        for c_idx, val in enumerate(values, start=1):
            cell = ws.cell(row=r_idx, column=c_idx, value=val)
            cell.fill = fill
            cell.alignment = Alignment(wrap_text=True, vertical="top")

    _autofit(ws)
    ws.row_dimensions[1].height = 18


def _build_questions(ws, survey: SurveyModel) -> None:
    ws.title = "Questions"

    headers = ["#", "Label", "Type", "Title", "Rows", "Cond", "Optional"]
    _header_row(ws, headers)

    ws.freeze_panes = "A2"

    for r_idx, q in enumerate(survey.questions(), start=2):
        fill = _ALT_FILL if r_idx % 2 == 0 else None
        values = [
            q.position + 1,
            q.label,
            q.tag,
            q.title,
            len(q.rows),
            q.cond or "",
            "Yes" if q.optional else "",
        ]
        for c_idx, val in enumerate(values, start=1):
            cell = ws.cell(row=r_idx, column=c_idx, value=val)
            if fill:
                cell.fill = fill
            cell.alignment = Alignment(wrap_text=True, vertical="top")

    _autofit(ws)


# ── Public API ─────────────────────────────────────────────────────────────────


def write_report(
    path: Path,
    survey: SurveyModel,
    findings: list[Finding],
) -> None:
    """Write a QA report Excel file to *path*.

    Args:
        path:     Destination .xlsx file path.
        survey:   Parsed survey model (for the Questions sheet).
        findings: All findings from run_checks() + run_routing_checks().
    """
    wb = Workbook()
    # Remove the default empty sheet
    wb.remove(wb.active)  # type: ignore[arg-type]

    _build_summary(wb.create_sheet("Summary"), survey, findings)
    _build_findings(wb.create_sheet("Findings"), findings)
    _build_questions(wb.create_sheet("Questions"), survey)

    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(path))
