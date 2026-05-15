"""Tests for the Excel reporter."""

from __future__ import annotations

from pathlib import Path

import pytest
from openpyxl import load_workbook

from survey_qa.core.models import Finding, SurveyModel
from survey_qa.reporters.excel import write_report


# ── Fixtures ──────────────────────────────────────────────────────────────────


@pytest.fixture
def sample_findings() -> list[Finding]:
    return [
        Finding(check_id="Q-001", severity="error", question_label="Q1", message="Missing in questionnaire"),
        Finding(check_id="Q-002", severity="warning", question_label="Q2", message="Title mismatch", detail="expected 'Foo' got 'Bar'"),
        Finding(check_id="Q-003", severity="info", question_label="Q3", message="Type hint mismatch"),
    ]


# ── Write and reload ──────────────────────────────────────────────────────────


def test_report_creates_file(tmp_path, survey_model, sample_findings):
    out = tmp_path / "report.xlsx"
    write_report(out, survey_model, sample_findings)
    assert out.exists()
    assert out.stat().st_size > 0


def test_report_has_three_sheets(tmp_path, survey_model, sample_findings):
    out = tmp_path / "report.xlsx"
    write_report(out, survey_model, sample_findings)
    wb = load_workbook(out)
    assert set(wb.sheetnames) == {"Summary", "Findings", "Questions"}


def test_summary_sheet_contains_survey_label(tmp_path, survey_model, sample_findings):
    out = tmp_path / "report.xlsx"
    write_report(out, survey_model, sample_findings)
    wb = load_workbook(out)
    ws = wb["Summary"]
    values = [str(ws.cell(row=r, column=2).value or "") for r in range(1, 12)]
    assert any(survey_model.survey_label in v for v in values)


def test_summary_error_count(tmp_path, survey_model, sample_findings):
    out = tmp_path / "report.xlsx"
    write_report(out, survey_model, sample_findings)
    wb = load_workbook(out)
    ws = wb["Summary"]
    # Find the "Errors" row and check its value is 1
    for row in ws.iter_rows(values_only=True):
        if row[0] == "Errors":
            assert row[1] == 1
            break


def test_findings_sheet_row_count(tmp_path, survey_model, sample_findings):
    out = tmp_path / "report.xlsx"
    write_report(out, survey_model, sample_findings)
    wb = load_workbook(out)
    ws = wb["Findings"]
    # Row 1 is header; remaining rows are findings
    data_rows = [r for r in ws.iter_rows(min_row=2, values_only=True) if any(c for c in r)]
    assert len(data_rows) == len(sample_findings)


def test_findings_sorted_errors_first(tmp_path, survey_model, sample_findings):
    out = tmp_path / "report.xlsx"
    write_report(out, survey_model, sample_findings)
    wb = load_workbook(out)
    ws = wb["Findings"]
    # Column 2 is Severity
    severities = [ws.cell(row=r, column=2).value for r in range(2, 2 + len(sample_findings))]
    assert severities[0] == "ERROR"


def test_questions_sheet_has_all_questions(tmp_path, survey_model):
    out = tmp_path / "report.xlsx"
    write_report(out, survey_model, [])
    wb = load_workbook(out)
    ws = wb["Questions"]
    data_rows = [r for r in ws.iter_rows(min_row=2, values_only=True) if any(c for c in r)]
    assert len(data_rows) == len(survey_model.questions())


def test_report_creates_parent_dirs(tmp_path, survey_model):
    out = tmp_path / "nested" / "deep" / "report.xlsx"
    write_report(out, survey_model, [])
    assert out.exists()


def test_empty_findings_still_writes(tmp_path, survey_model):
    out = tmp_path / "empty.xlsx"
    write_report(out, survey_model, [])
    wb = load_workbook(out)
    ws = wb["Findings"]
    data_rows = [r for r in ws.iter_rows(min_row=2, values_only=True) if any(c for c in r)]
    assert len(data_rows) == 0
